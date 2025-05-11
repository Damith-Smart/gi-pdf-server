from flask import Flask, request, send_file
from openpyxl import load_workbook
import os

app = Flask(__name__)


@app.route("/generate-pdf", methods=["POST"])
def generate_pdf():
    data = request.json

    # Load Excel template
    wb = load_workbook("gi_template.xlsx")
    ws = wb.active

    # Fill basic user inputs (merged cells or plain values)
    ws["E5"] = data.get("date", "")
    ws["E6"] = data.get("blast_in_charge", "")
    ws["E7"] = data.get("driller", "")
    ws["E9"] = data.get("time", "")
    ws["E10"] = data.get("material_type", "")
    ws["E11"] = data.get("location", "")

    # Blasting design inputs
    ws["E12"] = data.get("no_of_holes", 0)
    ws["E13"] = data.get("hole_depth", 0.0)
    ws["E14"] = data.get("sub_holes", 0)
    ws["E15"] = data.get("sub_depth", 0.0)
    ws["E16"] = data.get("spacing", 0.0)
    ws["E17"] = data.get("burden", 0.0)
    ws["E18"] = data.get("density", 0.0)

    # ✅ Watergel per hole — convert from cartridges to kg (1 cartridge = 0.125 kg)
    try:
        cartridges = int(data.get("watergel_per_hole", 0))
        ws["L11"] = cartridges * 0.125
    except:
        ws["L11"] = 0.0

    # ✅ Ammonium Nitrate per hole (kg) — only the per hole value
    try:
        ws["L28"] = float(data.get("ammonium_per_hole", 0))
    except:
        ws["L28"] = 0.0

    # ✅ ED Pattern (delays) — must be integers only, no strings like "number 01: x"
    ed_pattern = data.get("ed_pattern", [])
    for i in range(min(10, len(ed_pattern))):
        try:
            ws[f"L{15+i}"] = int(ed_pattern[i])
        except:
            ws[f"L{15+i}"] = ""

    # Save filled Excel
    filled_path = "filled_gi_form.xlsx"
    pdf_path = "gi_form.pdf"
    wb.save(filled_path)

    # Convert to PDF using LibreOffice
    os.system(
        f"libreoffice --headless --convert-to pdf {filled_path} --outdir .")

    if os.path.exists(pdf_path):
        return send_file(pdf_path, as_attachment=True)
    else:
        return {"error": "PDF generation failed"}, 500


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=10000)
